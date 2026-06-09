#!/usr/bin/env python3
"""
Import suppliers from Excel into riviera-marine production DB.

Reads 'Suppliers All over. (1).xlsx', extracts unique suppliers across all
country sheets, and generates SQL INSERT statements for:
  1. counterparties (SUPPLIER type)
  2. company_place_supply_rules (coverage per country)

Skips suppliers that already exist in the target tenant.

Usage:
  # Dry-run: just print what would be inserted
  python3 scripts/import_suppliers.py --dry-run

  # Generate SQL (pipe to psql or save to file)
  python3 scripts/import_suppliers.py > import_suppliers.sql

  # Execute directly against riviera-marine
  python3 scripts/import_suppliers.py | ssh root@139.162.157.31 "sudo -u postgres psql -d fueld"
"""

import argparse
import json
import os
import re
import sys
import uuid
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── Config ──────────────────────────────────────────────────────────────
EXCEL_PATH = Path(__file__).resolve().parent.parent / "uploads" / "Suppliers All over. (1).xlsx"
SKIP_SHEETS = {"Mapping"}

# riviera-marine tenant ID
TENANT_ID = os.environ.get("FUELD_TENANT_ID", "")

# ── Country name → ISO-3 code mapping ──────────────────────────────────
COUNTRY_ISO_MAP = {
    "albania": "ALB",
    "algeria": "DZA",
    "angola": "AGO",
    "argentina": "ARG",
    "australia": "AUS",
    "bangladesh": "BGD",
    "belgium": "BEL",
    "benin": "BEN",
    "brazil": "BRA",
    "bulgaria": "BGR",
    "cameroon": "CMR",
    "canada": "CAN",
    "cape verde": "CPV",
    "caribbean basin": "CAR",  # Not a country — use region code
    "chille": "CHL",
    "china": "CHN",
    "columbia": "COL",
    "congo": "COG",
    "croatia": "HRV",
    "cyprus": "CYP",
    "denmark": "DNK",
    "djibouti": "DJI",
    "dominican republic": "DOM",
    "egypt": "EGY",
    "english channel": "ENG",  # Region code
    "equatorial guinea": "GNQ",
    "estonia": "EST",
    "finland": "FIN",
    "france": "FRA",
    "gabon": "GAB",
    "gambia": "GMB",
    "germany": "DEU",
    "ghana": "GHA",
    "greece": "GRC",
    "guinea": "GIN",
    "guinea bissau": "GNB",
    "iceland": "ISL",
    "india": "IND",
    "indonesia": "IDN",
    "irland": "IRL",
    "israel": "ISR",
    "italy": "ITA",
    "ivory coast": "CIV",
    "jamaica": "JAM",
    "kenya": "KEN",
    "korea": "KOR",
    "latvia": "LVA",
    "liberia": "LBR",
    "libya": "LBY",
    "madagascar": "MDG",
    "malaysia": "MYS",
    "malta": "MLT",
    "mauritania": "MRT",
    "mautitius": "MUS",
    "montenegro": "MNE",
    "morocco": "MAR",
    "mozambique": "MOZ",
    "namibia": "NAM",
    "netherlands": "NLD",
    "new zealand": "NZL",
    "nigeria": "NGA",
    "norway": "NOR",
    "offshore west africa": "WAF",  # Region code
    "oman": "OMN",
    "pakistan": "PAK",
    "panama": "PAN",
    "philiphines": "PHL",
    "poland": "POL",
    "portugal": "PRT",
    "reunion": "REU",
    "romania": "ROU",
    "saudi arabia": "SAU",
    "sao tome and principe": "STP",
    "senegal": "SEN",
    "seychelles": "SYC",
    "sierra leone": "SLE",
    "singapore": "SGP",
    "somalia": "SOM",
    "south africa": "ZAF",
    "spain": "ESP",
    "sri lanka": "LKA",
    "sweden": "SWE",
    "tanzania": "TZA",
    "togo": "TGO",
    "trinidad": "TTO",
    "tunisia": "TUN",
    "turkey": "TUR",
    "uae": "ARE",
    "uae truck": "ARE",  # Same country, different delivery method
    "uk": "GBR",
    "usa": "USA",
    "uruguay": "URY",
}


def country_to_iso(name: str) -> str | None:
    """Map a sheet/country name to ISO-3 code."""
    return COUNTRY_ISO_MAP.get(name.strip().lower())


# ── Product parsing ────────────────────────────────────────────────────

def parse_products(grades_str: str) -> list[str]:
    """Parse 'Grades available' column into product type array.
    e.g. 'VLSFO/MGO' → ['VLSFO', 'MGO']
         'HSFO/VLSFO/MGO' → ['HSFO', 'VLSFO', 'MGO']
    """
    if not grades_str:
        return []
    # Split on / or , or space
    parts = re.split(r"[/,]\s*", grades_str)
    products = []
    for p in parts:
        p = p.strip().upper()
        if not p:
            continue
        # Normalize common variants
        if p in ("VSLFO",):
            p = "VLSFO"
        if p in ("LSMGO",):
            p = "LSMGO"
        products.append(p)
    return products


# ── Helpers ─────────────────────────────────────────────────────────────

def normalize_name(name: str) -> str:
    """Normalize a supplier name for dedup comparison."""
    return re.sub(r"\s+", " ", str(name).strip()).lower()


def clean_phone(phone) -> str:
    """Clean and normalize phone numbers."""
    if phone is None:
        return ""
    p = str(phone).strip()
    p = re.sub(r"^00\s*", "+", p)
    return p


def is_supplier_row(row: tuple, supplier_col: int = 0) -> bool:
    """Check if this row represents a new supplier (has a name in the supplier column)."""
    if supplier_col >= len(row):
        return False
    val = row[supplier_col]
    if val is None:
        return False
    name = str(val).strip()
    if not name:
        return False
    name_lower = name.lower()

    if name_lower == "none":
        return False
    if name_lower in {"supplier", "tel", "email", "port"}:
        return False
    if len(name) > 80:
        return False

    skip_keywords = {
        "disclaimer", "still an issue", "+ mma", "+ 550", "+ $", "if vsl",
        "saras also", "physical suppliers", "no limits", "delivery only",
        "currently all", "deliveries at", "all offers", "tankers are",
        "truck capacities", "some ports", "suppliers reverse",
        "all but minerva", "cepsa and minerva", "tenerife is",
        "galp considering", "bencom are", "repsol are presented",
        "repsol are stronger", "cepsa better", "marinoil often",
        "sinopec and chimbusco", "nunchi are not",
        "petromoc can", "12.6m", "10-15k",
        "spain trucks", "taragona - indulged",
    }
    for kw in skip_keywords:
        if name_lower.startswith(kw.lower()):
            return False

    return True


def extract_suppliers(excel_path: str) -> list[dict]:
    """Extract all unique suppliers from the Excel workbook."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    suppliers: dict[str, dict] = {}  # normalized_name -> supplier dict

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

        col_map = {}
        for idx, h in enumerate(headers):
            h_clean = h.rstrip()
            if "supplier" in h_clean and "areas" not in h_clean:
                col_map.setdefault("supplier", idx)
            elif "tel" in h_clean:
                col_map.setdefault("tel", idx)
            elif "email" in h_clean:
                col_map.setdefault("email", idx)
            elif "method" in h_clean or "delivery" in h_clean:
                col_map.setdefault("delivery", idx)
            elif "grade" in h_clean:
                col_map.setdefault("grades", idx)
            elif "payment" in h_clean:
                col_map.setdefault("payment_terms", idx)
            elif "claim" in h_clean:
                col_map.setdefault("claim_period", idx)
            elif "note" in h_clean:
                col_map.setdefault("notes", idx)
            elif "area" in h_clean:
                col_map.setdefault("areas", idx)

        supplier_col = col_map.get("supplier", 0)
        tel_col = col_map.get("tel")
        email_col = col_map.get("email")
        delivery_col = col_map.get("delivery")
        grades_col = col_map.get("grades")
        payment_col = col_map.get("payment_terms")
        claim_col = col_map.get("claim_period")
        notes_col = col_map.get("notes")
        areas_col = col_map.get("areas")

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not is_supplier_row(row, supplier_col):
                continue

            name = str(row[supplier_col]).strip()
            norm = normalize_name(name)

            if norm in suppliers:
                existing = suppliers[norm]
                if sheet_name not in existing["countries"]:
                    existing["countries"].append(sheet_name)
                continue

            phone = clean_phone(row[tel_col]) if tel_col is not None and tel_col < len(row) else ""
            email = str(row[email_col]).strip() if email_col is not None and email_col < len(row) and row[email_col] else ""
            delivery = str(row[delivery_col]).strip() if delivery_col is not None and delivery_col < len(row) and row[delivery_col] else ""
            grades = str(row[grades_col]).strip() if grades_col is not None and grades_col < len(row) and row[grades_col] else ""
            payment = str(row[payment_col]).strip() if payment_col is not None and payment_col < len(row) and row[payment_col] else ""
            claim = str(row[claim_col]).strip() if claim_col is not None and claim_col < len(row) and row[claim_col] else ""
            notes = str(row[notes_col]).strip() if notes_col is not None and notes_col < len(row) and row[notes_col] else ""
            areas = str(row[areas_col]).strip() if areas_col is not None and areas_col < len(row) and row[areas_col] else ""

            email = re.sub(r"^<|>$", "", email).strip()

            suppliers[norm] = {
                "name": name,
                "countries": [sheet_name],
                "phone": phone,
                "email": email,
                "delivery": delivery,
                "grades": grades,
                "payment_terms": payment,
                "claim_period": claim,
                "notes": notes,
                "areas": areas,
            }

    wb.close()
    return list(suppliers.values())


def escape_sql(s: str) -> str:
    """Escape single quotes for SQL."""
    return s.replace("'", "''")


def generate_sql(new_suppliers: list[dict], all_supplier_ids: dict[str, str], tenant_id: str) -> str:
    """Generate SQL INSERT statements for counterparties AND coverage rules."""
    lines = []
    lines.append("BEGIN;")
    lines.append("")
    lines.append(f"-- Import {len(new_suppliers)} new suppliers + coverage rules for all {len(all_supplier_ids)} suppliers")
    lines.append("")

    # ── 1. Counterparties (only NEW suppliers) ─────────────────────
    lines.append("-- ============================================================")
    lines.append("-- 1. COUNTERPARTIES (new suppliers only)")
    lines.append("-- ============================================================")
    lines.append("")

    for s in new_suppliers:
        sid = all_supplier_ids[normalize_name(s["name"])]
        name = escape_sql(s["name"])
        phone = escape_sql(s["phone"])
        email = escape_sql(s["email"])

        notes_parts = []
        if s["delivery"]:
            notes_parts.append(f"Delivery: {s['delivery']}")
        if s["grades"]:
            notes_parts.append(f"Grades: {s['grades']}")
        if s["payment_terms"]:
            notes_parts.append(f"Payment: {s['payment_terms']}")
        if s["claim_period"]:
            notes_parts.append(f"Claim: {s['claim_period']}")
        if s["areas"]:
            notes_parts.append(f"Areas: {s['areas']}")
        if s["notes"]:
            notes_parts.append(s["notes"])

        country = s["countries"][0] if s["countries"] else None
        country_sql = f"'{escape_sql(country)}'" if country else "NULL"

        segments = {}
        if len(s["countries"]) > 1:
            segments["countries"] = s["countries"]
        if notes_parts:
            segments["import_notes"] = " | ".join(notes_parts)

        segments_sql = "NULL"
        if segments:
            segments_sql = f"'{escape_sql(json.dumps(segments))}'"

        phone_sql = "NULL" if not phone else f"'{phone}'"
        email_sql = "NULL" if not email else f"'{email}'"

        lines.append(
            f"INSERT INTO counterparties (id, tenant_id, name, type, country, head_office_phone, head_office_email, segments, created_at, updated_at) "
            f"VALUES ('{sid}', '{tenant_id}', '{name}', 'SUPPLIER', {country_sql}, "
            f"{phone_sql}, {email_sql}, {segments_sql}, NOW(), NOW());"
        )

    # ── 2. Coverage rules (ALL suppliers from Excel) ─────────────────
    lines.append("")
    lines.append("-- ============================================================")
    lines.append("-- 2. COMPANY PLACE SUPPLY RULES (coverage per country)")
    lines.append("-- ============================================================")
    lines.append("")

    # We need ALL suppliers from the Excel to generate coverage rules
    # Re-extract to get the full list
    all_suppliers = extract_suppliers(str(EXCEL_PATH))

    coverage_count = 0
    for s in all_suppliers:
        norm = normalize_name(s["name"])
        if norm not in all_supplier_ids:
            continue  # Shouldn't happen, but safety check
        sid = all_supplier_ids[norm]
        products = parse_products(s["grades"])

        for country_name in s["countries"]:
            iso = country_to_iso(country_name)
            if not iso:
                print(f"WARNING: No ISO code for country '{country_name}' — skipping coverage rule for {s['name']}", file=sys.stderr)
                continue

            rid = str(uuid.uuid4())
            products_json = json.dumps(products)
            products_sql = f"'{escape_sql(products_json)}'"

            # Build note from delivery method + payment terms
            note_parts = []
            if s["delivery"]:
                note_parts.append(f"Delivery: {s['delivery']}")
            if s["payment_terms"]:
                note_parts.append(f"Payment: {s['payment_terms']}")
            note_sql = "NULL"
            if note_parts:
                note_sql = f"'{escape_sql(' | '.join(note_parts))}'"

            lines.append(
                f"INSERT INTO company_place_supply_rules (id, company_id, country_iso, place_types, products, note, is_active, created_at, updated_at) "
                f"VALUES ('{rid}', '{sid}', '{iso}', '[]', {products_sql}, {note_sql}, true, NOW(), NOW());"
            )
            coverage_count += 1

    lines.append("")
    lines.append("COMMIT;")

    # Summary to stderr
    print(f"-- {len(new_suppliers)} new suppliers, {coverage_count} coverage rules generated", file=sys.stderr)

    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="Import suppliers from Excel")
    parser.add_argument("--dry-run", action="store_true", help="Print summary only, no SQL")
    parser.add_argument("--tenant-id", default=TENANT_ID, help="Tenant UUID (or set FUELD_TENANT_ID env var)")
    parser.add_argument("--existing", help="File with existing supplier names (one per line) for dedup")
    parser.add_argument("--existing-with-ids", help="File with existing supplier IDs: 'id | name' format. Coverage rules will be created for ALL suppliers.")
    args = parser.parse_args()

    if not args.tenant_id:
        print("ERROR: --tenant-id is required (or set FUELD_TENANT_ID)", file=sys.stderr)
        sys.exit(1)

    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}", file=sys.stderr)
        sys.exit(1)

    suppliers = extract_suppliers(str(EXCEL_PATH))

    # Load existing names if provided
    existing_names = set()
    if args.existing:
        with open(args.existing) as f:
            existing_names = {normalize_name(line) for line in f if line.strip()}

    # Load existing suppliers with IDs for coverage rule generation
    existing_supplier_ids: dict[str, str] = {}  # normalized_name -> id
    if args.existing_with_ids:
        with open(args.existing_with_ids) as f:
            for line in f:
                line = line.strip()
                if not line or "|" not in line:
                    continue
                parts = line.split("|", 1)
                if len(parts) == 2:
                    sid = parts[0].strip()
                    sname = parts[1].strip()
                    existing_supplier_ids[normalize_name(sname)] = sid

    # Filter out existing for counterparty creation
    new_suppliers = [s for s in suppliers if normalize_name(s["name"]) not in existing_names]
    skipped = len(suppliers) - len(new_suppliers)

    # Build full mapping: new suppliers get fresh UUIDs, existing use their DB IDs
    all_supplier_ids: dict[str, str] = {}
    for s in suppliers:
        norm = normalize_name(s["name"])
        if norm in existing_supplier_ids:
            all_supplier_ids[norm] = existing_supplier_ids[norm]
        else:
            all_supplier_ids[norm] = str(uuid.uuid4())

    if args.dry_run:
        print(f"Total unique suppliers in Excel: {len(suppliers)}")
        print(f"Already in DB (counterparty skipped): {skipped}")
        print(f"New suppliers to create: {len(new_suppliers)}")
        print(f"Existing suppliers getting coverage rules: {len(existing_supplier_ids)}")
        print()

        # Count coverage rules for ALL suppliers
        total_rules = 0
        unmapped = set()
        for s in suppliers:
            for c in s["countries"]:
                iso = country_to_iso(c)
                if iso:
                    total_rules += 1
                else:
                    unmapped.add(c)

        print(f"Coverage rules to create (all suppliers): {total_rules}")
        if unmapped:
            print(f"Unmapped countries (no ISO code): {sorted(unmapped)}")
        print()

        print("New suppliers:")
        for s in new_suppliers:
            countries = ", ".join(s["countries"])
            products = parse_products(s["grades"])
            print(f"  {s['name']} ({countries})")
            if s["phone"]:
                print(f"    📞 {s['phone']}")
            if s["email"]:
                print(f"    ✉️  {s['email']}")
            if products:
                print(f"    🛢️  {', '.join(products)}")
            if s["delivery"]:
                print(f"    🚢 {s['delivery']}")
        return

    sql = generate_sql(new_suppliers, all_supplier_ids, args.tenant_id)
    print(sql)


if __name__ == "__main__":
    main()


if __name__ == "__main__":
    main()
